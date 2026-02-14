import tkinter as tk
from tkinter import messagebox, filedialog
from datetime import datetime, timedelta
from openpyxl import load_workbook

# name you excel spread sheet here
TIMESHEET_FILE = ""

today_date = datetime.now()


class FlexiTimeSheet:
    def __init__(self, root):
        self.root = root
        self.root.title("Flexi TimeSheet")
        self.root.geometry("285x220")

        self.clock_in_time = None
        self.break_start_time = None
        self.break_end_time = None
        self.clock_out_time = None

        self.datetime_label = tk.Label(root, text="")
        self.datetime_label.pack(pady=5)
        self.update_datetime()

        tk.Button(root, text="Clock In", width=20, command=self.clock_in).pack(pady=5)
        tk.Button(root, text="Start Break", width=20, command=self.start_break).pack(pady=5)
        tk.Button(root, text="End Break", width=20, command=self.end_break).pack(pady=5)
        tk.Button(root, text="Clock Out", width=20, command=self.clock_out).pack(pady=5)

        self.status = tk.Label(root, text="Not clocked in", fg="red")
        self.status.pack(pady=10)

    def update_datetime(self):
        now = datetime.now()
        self.datetime_label.config(text=now.strftime("%A, %d %B %Y %H:%M:%S"))
        self.root.after(1000, self.update_datetime)


    def clock_in(self):
        wb = load_workbook(TIMESHEET_FILE)
        ws = wb.active
        today_day = today_date.day
        target_date = None

        for row in range(11, 44):
            if ws[f"B{row}"].value == today_day:
                target_date = row
                break

        if target_date is None:
            messagebox.showwarning("Error",f"'{today_date.strftime('%d/%m/%Y')}' is not in your Time Sheet!")
        else:
            self.clock_in_time = datetime.now().time()
            self.status.config(text="Clocked In", fg="green")


    def start_break(self):
        if self.clock_in_time:
            # simulate break at 12:30
            self.break_start_time = datetime.now().time()
            self.status.config(text="Started break", fg="green")
        else:
            messagebox.showwarning("Error", "Clock in first!")


    def end_break(self):
        if self.break_start_time:
            # simulate break end at 13:00
            self.break_end_time = datetime.now().time()
            self.status.config(text="Ended break", fg="green")
        else:
            messagebox.showwarning("Error", "Start break first!")


    def clock_out(self):
        if not self.clock_in_time:
            messagebox.showwarning("Error", "Clock in first!")
            return

        # simulate clock out at 17:00
        self.clock_out_time = datetime.now().time()
        self.status.config(text="Clocked Out", fg="green")

        try:
            self.write_to_excel()
            messagebox.showinfo("Success", "This has now been saved to your TimeSheet file")
            self.reset()
        except Exception as e:
            messagebox.showerror("Error", str(e))


    def write_to_excel(self):
        wb = load_workbook(TIMESHEET_FILE)
        ws = wb.active

        today_day = today_date.day

        target_date = None

        for row in range(11, 44):
            if ws[f"B{row}"].value == today_day:
                target_date = row
                break

        if target_date is None:
            raise Exception(f"'{today_date.strftime("%d/%m/%Y")}' is not in your Time Sheet!")

        # Write times
        ws[f"D{target_date}"] = self.clock_in_time
        ws[f"E{target_date}"] = self.break_start_time
        ws[f"F{target_date}"] = self.break_end_time
        ws[f"G{target_date}"] = self.clock_out_time

        wb.save(TIMESHEET_FILE)


    def reset(self):
        self.clock_in_time = None
        self.break_start_time = None
        self.break_end_time = None
        self.clock_out_time = None
        self.status.config(text="You are currently not clocked in", fg="red")


root = tk.Tk()
app = FlexiTimeSheet(root)
root.mainloop()
